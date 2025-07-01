import streamlit as st
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os
from datetime import datetime, timedelta
import warnings
import io
import tempfile
import zipfile
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
warnings.filterwarnings('ignore')

# Configuração da página
st.set_page_config(
    page_title="Medições Usina Geradora Floriano",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="expanded"
)

# CSS personalizado com as cores da CSN
st.markdown("""
<style>
    .main-header {
        background: linear-gradient(90deg, #00529C 0%, #231F20 100%);
        padding: 2rem;
        border-radius: 15px;
        margin-bottom: 2rem;
        text-align: center;
        box-shadow: 0 4px 15px rgba(0,0,0,0.2);
        position: relative;
    }
    
    .logo-container {
        display: flex;
        align-items: center;
        justify-content: center;
        gap: 2rem;
        margin-bottom: 1rem;
    }
    
    .logo-img {
        height: 80px;
        width: auto;
        background: white;
        padding: 10px;
        border-radius: 10px;
        box-shadow: 0 2px 10px rgba(0,0,0,0.3);
    }
    
    .header-text {
        flex: 1;
        text-align: center;
    }
    
    .main-header h1 {
        color: white;
        font-size: 2.8rem;
        margin: 0;
        font-weight: bold;
        text-shadow: 2px 2px 4px rgba(0,0,0,0.3);
    }
    
    .main-header p {
        color: #E8E8E8;
        font-size: 1.3rem;
        margin: 0.5rem 0 0 0;
    }
    
    .stButton > button {
        background: linear-gradient(45deg, #00529C, #0066CC);
        color: white;
        border: none;
        border-radius: 8px;
        padding: 0.7rem 1.5rem;
        font-weight: bold;
        font-size: 1.1rem;
        transition: all 0.3s ease;
        box-shadow: 0 3px 10px rgba(0,82,156,0.3);
    }
    
    .stButton > button:hover {
        background: linear-gradient(45deg, #231F20, #404040);
        transform: translateY(-2px);
        box-shadow: 0 5px 15px rgba(0,82,156,0.4);
    }
    
    .success-box {
        background: linear-gradient(135deg, #d4edda, #c3e6cb);
        border: 1px solid #c3e6cb;
        border-radius: 10px;
        padding: 1.5rem;
        margin: 1rem 0;
        box-shadow: 0 2px 8px rgba(0,0,0,0.1);
    }
    
    .error-box {
        background: linear-gradient(135deg, #f8d7da, #f5c6cb);
        border: 1px solid #f5c6cb;
        border-radius: 10px;
        padding: 1.5rem;
        margin: 1rem 0;
        box-shadow: 0 2px 8px rgba(0,0,0,0.1);
    }
    
    .info-box {
        background: linear-gradient(135deg, #d1ecf1, #bee5eb);
        border: 1px solid #bee5eb;
        border-radius: 10px;
        padding: 1.5rem;
        margin: 1rem 0;
        box-shadow: 0 2px 8px rgba(0,0,0,0.1);
    }
    
    .warning-box {
        background: linear-gradient(135deg, #fff3cd, #ffeaa7);
        border: 1px solid #ffeaa7;
        border-radius: 10px;
        padding: 1.5rem;
        margin: 1rem 0;
        box-shadow: 0 2px 8px rgba(0,0,0,0.1);
    }
    
    .metric-card {
        background: linear-gradient(135deg, #ffffff, #f8f9fa);
        padding: 1.5rem;
        border-radius: 12px;
        border-left: 5px solid #00529C;
        margin: 0.5rem 0;
        box-shadow: 0 4px 12px rgba(0,0,0,0.1);
        transition: transform 0.2s ease;
    }
    
    .metric-card:hover {
        transform: translateY(-2px);
        box-shadow: 0 6px 20px rgba(0,0,0,0.15);
    }
    
    .metric-card h4 {
        color: #00529C;
        margin: 0 0 0.5rem 0;
        font-size: 1rem;
        font-weight: 600;
    }
    
    .metric-card h2 {
        color: #231F20;
        margin: 0;
        font-size: 2rem;
        font-weight: bold;
    }
    
    .stats-container {
        background: linear-gradient(135deg, #f8f9fa, #e9ecef);
        padding: 2rem;
        border-radius: 15px;
        margin: 1rem 0;
        box-shadow: 0 4px 15px rgba(0,0,0,0.1);
    }
    
    .chart-container {
        background: white;
        padding: 1.5rem;
        border-radius: 12px;
        margin: 1rem 0;
        box-shadow: 0 3px 10px rgba(0,0,0,0.1);
        border: 1px solid #e9ecef;
    }
    
    .dashboard-header {
        background: linear-gradient(135deg, #00529C, #0066CC);
        color: white;
        padding: 1.5rem;
        border-radius: 12px;
        margin: 1rem 0 2rem 0;
        text-align: center;
        box-shadow: 0 4px 15px rgba(0,82,156,0.3);
    }
    
    .dashboard-filters {
        background: linear-gradient(135deg, #f8f9fa, #ffffff);
        padding: 1.5rem;
        border-radius: 12px;
        margin: 1rem 0;
        border: 1px solid #e9ecef;
        box-shadow: 0 2px 8px rgba(0,0,0,0.05);
    }
    
    .sidebar .sidebar-content {
        background: linear-gradient(180deg, #f8f9fa, #ffffff);
    }
    
    .stDataFrame {
        border-radius: 10px;
        overflow: hidden;
        box-shadow: 0 2px 8px rgba(0,0,0,0.1);
    }
    
    .stExpander {
        border-radius: 10px;
        border: 1px solid #e9ecef;
        box-shadow: 0 2px 8px rgba(0,0,0,0.05);
    }
    
    /* Responsividade para dispositivos móveis */
    @media (max-width: 768px) {
        .logo-container {
            flex-direction: column;
            gap: 1rem;
        }
        
        .main-header h1 {
            font-size: 2.2rem;
        }
        
        .logo-img {
            height: 60px;
        }
    }
</style>
""", unsafe_allow_html=True)

class ExactWeatherProcessor:
    """
    Processador de dados meteorológicos com busca EXATA
    NÃO faz médias ou inferências - apenas busca dados pontuais com tolerância de ±10 minutos
    """
    def __init__(self):
        self.consolidated_data = {}  # {timestamp: {variavel: valor}}
        self.processed_sheets = []
        self.conflicts_detected = []
        self.excel_path = None
        
        # Mapeamento de colunas para análise diária
        self.column_mapping = {
            'Temperatura': {'start_num': 2},        # B até AF (2-32)
            'Piranometro_1': {'start_num': 33},     # AG até BK (33-63)
            'Piranometro_2': {'start_num': 64},     # BL até CP (64-94)
            'Piranometro_Alab': {'start_num': 95},  # CQ até DU (95-125)
            'Umidade_Relativa': {'start_num': 126}, # DV até EZ (126-156)
            'Velocidade_Vento': {'start_num': 157}  # FA até GE (157-187)
        }

        # Mapeamento de colunas para análise mensal
        self.monthly_column_mapping = {
            'Temperatura': {'start_col': 'B', 'rows': (3, 33)},      # B3:E33
            'Piranometro_1': {'start_col': 'H', 'rows': (3, 33)},   # H3:K33  
            'Piranometro_2': {'start_col': 'N', 'rows': (3, 33)},   # N3:Q33
            'Piranometro_Alab': {'start_col': 'T', 'rows': (3, 33)}, # T3:W33
            'Umidade_Relativa': {'start_col': 'Z', 'rows': (3, 33)}, # Z3:AC33
            'Velocidade_Vento': {'start_col': 'B', 'rows': (37, 67)}, # B37:E67
            'Bateria': {'start_col': 'H', 'rows': (37, 67)},        # H37:K67
            'LitBatt': {'start_col': 'N', 'rows': (37, 67)},        # N37:Q67
            'LogTemp': {'start_col': 'T', 'rows': (37, 67)}         # T37:W67
        }

        # Mapeamento .dat → análise mensal (incluindo variáveis adicionais)
        self.monthly_variable_mapping = {
            'Temp_Avg': 'Temperatura',
            'Pir1_Avg': 'Piranometro_1', 
            'Pir2_Avg': 'Piranometro_2',
            'PirALB_Avg': 'Piranometro_Alab',
            'RH_Avg': 'Umidade_Relativa',
            'Ane_Avg': 'Velocidade_Vento',
            'Batt_Avg': 'Bateria',
            'LitBatt_Avg': 'LitBatt',
            'LoggTemp_Avg': 'LogTemp'
        }

    def process_dat_files(self, dat_files):
        """Processa múltiplos arquivos .dat consolidando por TIMESTAMP exato"""
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        total_files = len(dat_files)
        self.file_processing_info = []
        self.conflicts_detected = []
        
        # ETAPA 1: Ler todos os arquivos e consolidar
        for i, uploaded_file in enumerate(dat_files):
            status_text.text(f"Processando {i+1}/{total_files}: {uploaded_file.name}")
            try:
                uploaded_file.seek(0)
                data = pd.read_csv(uploaded_file, skiprows=4, parse_dates=[0])
                data.columns = [
                    'TIMESTAMP', 'RECORD',
                    'Ane_Min', 'Ane_Max', 'Ane_Avg', 'Ane_Std',
                    'Temp_Min', 'Temp_Max', 'Temp_Avg', 'Temp_Std',
                    'RH_Min', 'RH_Max', 'RH_Avg', 'RH_Std',
                    'Pir1_Min', 'Pir1_Max', 'Pir1_Avg', 'Pir1_Std',
                    'Pir2_Min', 'Pir2_Max', 'Pir2_Avg', 'Pir2_Std',
                    'PirALB_Min', 'PirALB_Max', 'PirALB_Avg', 'PirALB_Std',
                    'Batt_Min', 'Batt_Max', 'Batt_Avg', 'Batt_Std',
                    'LoggTemp_Min', 'LoggTemp_Max', 'LoggTemp_Avg', 'LoggTemp_Std',
                    'LitBatt_Min', 'LitBatt_Max', 'LitBatt_Avg', 'LitBatt_Std'
                ]
                
                # Consolidar dados timestamp por timestamp
                for _, row in data.iterrows():
                    timestamp = row['TIMESTAMP']
                                 
                    # Extrair dados das variáveis (incluindo novas para análise mensal)
                    new_data = {
                        'Temperatura': round(row['Temp_Avg'], 2) if not pd.isna(row['Temp_Avg']) else None,
                        'Piranometro_1': round(row['Pir1_Avg'] / 1000, 3) if not pd.isna(row['Pir1_Avg']) else None,
                        'Piranometro_2': round(row['Pir2_Avg'] / 1000, 3) if not pd.isna(row['Pir2_Avg']) else None,
                        'Piranometro_Alab': round(row['PirALB_Avg'] / 1000, 3) if not pd.isna(row['PirALB_Avg']) else None,
                        'Umidade_Relativa': round(row['RH_Avg'], 2) if not pd.isna(row['RH_Avg']) else None,
                        'Velocidade_Vento': round(row['Ane_Avg'], 2) if not pd.isna(row['Ane_Avg']) else None,
                        # Novas variáveis para análise mensal
                        'Bateria': round(row['Batt_Avg'], 2) if not pd.isna(row['Batt_Avg']) else None,
                        'LitBatt': round(row['LitBatt_Avg'], 2) if not pd.isna(row['LitBatt_Avg']) else None,
                        'LogTemp': round(row['LoggTemp_Avg'], 2) if not pd.isna(row['LoggTemp_Avg']) else None
                    }
                    
                    # Verificar se já existe dados para este timestamp
                    if timestamp in self.consolidated_data:
                        # CONFLITO DETECTADO!
                        conflict_info = {
                            'timestamp': timestamp,
                            'arquivo_anterior': 'dados_anteriores',
                            'arquivo_atual': uploaded_file.name,
                            'dados_anteriores': self.consolidated_data[timestamp].copy(),
                            'dados_novos': new_data.copy()
                        }
                        self.conflicts_detected.append(conflict_info)
                        
                        # Usar último arquivo (sobrescrever)
                        self.consolidated_data[timestamp] = new_data
                    else:
                        # Novo timestamp, adicionar
                        self.consolidated_data[timestamp] = new_data
                
                # Info do arquivo processado
                self.file_processing_info.append({
                    'arquivo': uploaded_file.name,
                    'registros': len(data),
                    'periodo_inicio': data['TIMESTAMP'].min().strftime('%Y-%m-%d %H:%M'),
                    'periodo_fim': data['TIMESTAMP'].max().strftime('%Y-%m-%d %H:%M'),
                    'status': 'Processado com sucesso'
                })
                
            except Exception as e:
                self.file_processing_info.append({
                    'arquivo': uploaded_file.name,
                    'registros': 0,
                    'periodo_inicio': 'N/A',
                    'periodo_fim': 'N/A',
                    'status': f'Erro: {str(e)}'
                })
            
            progress_bar.progress((i + 1) / total_files)
        
        status_text.text("Consolidação concluída com sucesso!")
        
        # Mostrar conflitos se detectados
        if self.conflicts_detected:
            self._show_conflicts()
        
        # Mostrar resumo do processamento
        self._show_file_processing_summary()
        
        return len(self.consolidated_data) > 0

    def _show_conflicts(self):
        """Mostra conflitos detectados entre arquivos"""
        st.markdown("---")
        st.markdown("### Conflitos Detectados")
        
        st.markdown(f"""
        <div class="warning-box">
            <h4>{len(self.conflicts_detected)} conflito(s) encontrado(s)</h4>
            <p>Timestamps idênticos em múltiplos arquivos. Usando dados do último arquivo processado.</p>
        </div>
        """, unsafe_allow_html=True)
        
        # Mostrar detalhes dos conflitos
        with st.expander("Ver Detalhes dos Conflitos"):
            for i, conflict in enumerate(self.conflicts_detected[:10]):  # Mostrar só os primeiros 10
                st.markdown(f"**Conflito {i+1}: {conflict['timestamp']}**")
                col1, col2 = st.columns(2)
                with col1:
                    st.markdown("*Dados Anteriores:*")
                    st.json(conflict['dados_anteriores'])
                with col2:
                    st.markdown(f"*Dados Novos ({conflict['arquivo_atual']}):*")
                    st.json(conflict['dados_novos'])
                st.markdown("---")
            
            if len(self.conflicts_detected) > 10:
                st.info(f"Mostrando apenas os primeiros 10 conflitos de {len(self.conflicts_detected)} total.")

    def _find_closest_timestamp(self, target_time, available_timestamps):
        """
        Busca o timestamp mais próximo dentro da tolerância de ±10 minutos
        
        Args:
            target_time: datetime alvo (ex: 2025-06-22 10:00:00)
            available_timestamps: lista de timestamps disponíveis
            
        Returns:
            timestamp mais próximo ou None se nenhum estiver dentro da tolerância
        """
        target_time = pd.to_datetime(target_time)
        tolerance = timedelta(minutes=10)
        
        min_diff = timedelta.max
        closest_timestamp = None
        
        for ts in available_timestamps:
            ts_converted = pd.to_datetime(ts)
            diff = abs(ts_converted - target_time)
            
            # Verifica se está dentro da tolerância e é mais próximo
            if diff <= tolerance and diff < min_diff:
                min_diff = diff
                closest_timestamp = ts  # Retorna o timestamp original, não o convertido
        
        return closest_timestamp

    def update_excel_file(self, excel_file):
        """
        Atualiza Excel com dados exatos
        """
        if not self.consolidated_data:
            return False, "Nenhum dado processado!"
        
        try:
            # Salvar arquivo Excel temporariamente
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
                tmp_file.write(excel_file.read())
                self.excel_path = tmp_file.name
            
            wb = load_workbook(self.excel_path)
            
            progress_bar = st.progress(0)
            status_text = st.empty()
            
            # Agrupar dados por mês
            monthly_data = {}
            for timestamp, data in self.consolidated_data.items():
                year_month = f"{timestamp.year}-{timestamp.month:02d}"
                if year_month not in monthly_data:
                    monthly_data[year_month] = {}
                monthly_data[year_month][timestamp] = data
            
            total_months = len(monthly_data)
            sheets_updated = 0
            total_cells_updated = 0
            
            # Processar cada mês - ANÁLISES DIÁRIAS
            for i, (year_month, month_timestamps) in enumerate(monthly_data.items()):
                year, month = year_month.split('-')
                month_num = int(month)
                
                status_text.text(f"Processando análise diária - mês {month}/{year}...")
                
                # Buscar aba correspondente
                sheet_name = self._find_daily_analysis_sheet(wb.sheetnames, month_num)
                if sheet_name:
                    ws = wb[sheet_name]
                    cells_updated = self._update_daily_analysis_exact(ws, month_timestamps, int(year), month_num)
                    
                    if cells_updated > 0:
                        sheets_updated += 1
                        total_cells_updated += cells_updated
                        self.processed_sheets.append(sheet_name)
                
                progress_bar.progress((i + 1) / (total_months * 2))  # Ajustar para incluir análise mensal
            
            # PROCESSAR ANÁLISES MENSAIS
            status_text.text("Processando análises mensais...")
            monthly_sheets_updated, monthly_cells_updated = self._process_monthly_analysis(wb, monthly_data)

            # Atualizar totais
            sheets_updated += monthly_sheets_updated
            total_cells_updated += monthly_cells_updated

            # Salvar alterações
            wb.save(self.excel_path)
            status_text.text("Atualização concluída com sucesso!")

            if sheets_updated > 0:
                diarias_msg = f"{sheets_updated - monthly_sheets_updated} aba(s) diária(s)" if sheets_updated > monthly_sheets_updated else ""
                mensais_msg = f"{monthly_sheets_updated} aba(s) mensal(is)" if monthly_sheets_updated > 0 else ""
                separador = " e " if diarias_msg and mensais_msg else ""
                
                return True, f"Sucesso! {diarias_msg}{separador}{mensais_msg} atualizada(s), {total_cells_updated} célula(s) preenchida(s)"
            else:
                return False, "Nenhuma aba compatível encontrada para atualização"
                
        except Exception as e:
            return False, f"Erro durante atualização: {e}"

    def _find_daily_analysis_sheet(self, sheet_names, month_num):
        """Encontra aba de análise diária para o mês"""
        month_str = f"{month_num:02d}"
        target_pattern = f"{month_str}-Analise Diaria"
        
        # Busca exata primeiro
        if target_pattern in sheet_names:
            return target_pattern
        
        # Busca por padrão similar
        for sheet_name in sheet_names:
            if month_str in sheet_name and "Analise Diaria" in sheet_name:
                return sheet_name
        
        return None

    def _update_daily_analysis_exact(self, ws, month_timestamps, year, month):
        """
        Atualiza análise diária usando busca exata
        """
        cells_updated = 0
        
        # Para cada horário da planilha (00:00 a 23:00)
        for hour in range(24):
            row_num = hour + 3  # Linha 3 = 00:00, Linha 4 = 01:00, etc.
            
            # Para cada dia do mês (1 a 31)
            for day in range(1, 32):
                # Construir timestamp alvo
                try:
                    target_datetime = datetime(year, month, day, hour, 0, 0)
                except ValueError:
                    # Dia inválido para o mês (ex: 31 de fevereiro)
                    continue
                
                # Buscar timestamp mais próximo dentro da tolerância
                available_timestamps = list(month_timestamps.keys())
                closest_timestamp = self._find_closest_timestamp(target_datetime, available_timestamps)
                
                if closest_timestamp is None:
                    # Nenhum dado dentro da tolerância - deixar vazio
                    continue
                
                # Obter dados do timestamp encontrado
                data = month_timestamps[closest_timestamp]
                
                # Atualizar cada variável
                for variable, value in data.items():
                    if value is None:
                        continue
                    
                    col_letter = self._get_column_for_variable_and_day(variable, day)
                    if col_letter:
                        try:
                            ws[f'{col_letter}{row_num}'] = value
                            cells_updated += 1
                        except Exception:
                            # Continua processamento mesmo com erro
                            pass
        
        return cells_updated

    def _get_column_for_variable_and_day(self, variable, day_number):
        """
        Calcula letra da coluna para análise diária
        """
        if variable not in self.column_mapping:
            return None
        
        # Verificar se o dia está no range válido (1-31)
        if day_number < 1 or day_number > 31:
            return None
        
        start_col_num = self.column_mapping[variable]['start_num']
        target_col_num = start_col_num + (day_number - 1)
        
        # Verificar se a coluna está dentro dos limites válidos
        if target_col_num > 187:  # Última coluna GE = 187
            return None
            
        col_letter = get_column_letter(target_col_num)
        return col_letter

    def _find_monthly_analysis_sheet(self, sheet_names, month_num):
        """Encontra aba de análise mensal para o mês"""
        month_str = f"{month_num:02d}"
        target_pattern = f"{month_str}-Analise Mensal"
        
        # Busca exata primeiro
        if target_pattern in sheet_names:
            return target_pattern
        
        # Busca por padrão similar
        for sheet_name in sheet_names:
            if month_str in sheet_name and "Analise Mensal" in sheet_name:
                return sheet_name
        
        return None

    def _process_monthly_analysis(self, wb, monthly_data):
        """Processa todas as abas de análise mensal - VERSÃO CORRIGIDA"""
        monthly_sheets_updated = 0
        monthly_cells_updated = 0
        
        print(f"🔍 DEBUG: Iniciando análise mensal...")
        print(f"🔍 DEBUG: Meses disponíveis: {list(monthly_data.keys())}")
        print(f"🔍 DEBUG: Abas no Excel: {wb.sheetnames}")
        
        # Debug do mapeamento de colunas
        self._debug_column_mapping()
        
        for year_month, month_timestamps in monthly_data.items():
            year, month = year_month.split('-')
            month_num = int(month)
            
            print(f"🔍 DEBUG: Processando {year_month} (mês {month_num})")
            print(f"🔍 DEBUG: Timestamps disponíveis: {len(month_timestamps)}")
            
            # Verificar variáveis disponíveis
            common_vars = self._verify_data_variables(month_timestamps)
            
            # Buscar aba mensal correspondente
            monthly_sheet_name = self._find_monthly_analysis_sheet(wb.sheetnames, month_num)
            print(f"🔍 DEBUG: Aba encontrada: {monthly_sheet_name}")
            
            if monthly_sheet_name:
                ws_monthly = wb[monthly_sheet_name]
                
                print(f"🔍 DEBUG: Iniciando processamento da aba {monthly_sheet_name}")
                
                # Debug adicional: verificar algumas células da planilha
                self._debug_worksheet_structure(ws_monthly)
                
                cells_updated = self._update_monthly_analysis_data(ws_monthly, month_timestamps, int(year), month_num)
                print(f"🔍 DEBUG: Células atualizadas na aba mensal: {cells_updated}")
                
                if cells_updated > 0:
                    monthly_sheets_updated += 1
                    monthly_cells_updated += cells_updated
                    self.processed_sheets.append(f"{monthly_sheet_name} (Mensal)")
            else:
                print(f"❌ DEBUG: Nenhuma aba mensal encontrada para mês {month_num}")
                print(f"❌ DEBUG: Procurando por: '{month_num:02d}-Analise Mensal'")
                
                # Mostrar abas similares para debug
                similar_sheets = [s for s in wb.sheetnames if str(month_num).zfill(2) in s and 'Mensal' in s]
                if similar_sheets:
                    print(f"❌ DEBUG: Abas similares encontradas: {similar_sheets}")
        
        print(f"🔍 DEBUG: RESULTADO FINAL - Abas mensais: {monthly_sheets_updated}, Células: {monthly_cells_updated}")
        return monthly_sheets_updated, monthly_cells_updated

    def _update_monthly_analysis_data(self, ws, month_timestamps, year, month):
        """Atualiza análise mensal com estatísticas diárias - VERSÃO CORRIGIDA"""
        cells_updated = 0
        
        print(f"🔍 DEBUG: Iniciando update da aba mensal para {month}/{year}")
        print(f"🔍 DEBUG: Total de timestamps disponíveis: {len(month_timestamps)}")
        
        # Verificar quais variáveis temos nos dados
        if not month_timestamps:
            print("❌ DEBUG: Nenhum timestamp disponível")
            return 0
            
        sample_data = next(iter(month_timestamps.values()))
        available_variables = list(sample_data.keys())
        print(f"🔍 DEBUG: Variáveis disponíveis nos dados: {available_variables}")
        
        # Verificar quais variáveis estão no mapeamento
        mapped_variables = list(self.monthly_column_mapping.keys())
        print(f"🔍 DEBUG: Variáveis no mapeamento: {mapped_variables}")
        
        # Para cada dia do mês (1 a 31)
        for day in range(1, 32):
            try:
                # Verificar se o dia existe no mês
                datetime(year, month, day)
            except ValueError:
                # Dia inválido para o mês (ex: 31 de fevereiro)
                continue
            
            # Filtrar todos os timestamps do dia
            day_timestamps = [ts for ts in month_timestamps.keys() if ts.day == day]
            
            if not day_timestamps:
                # Não há dados para este dia - deixar células vazias
                continue
            
            print(f"🔍 DEBUG: Dia {day} - {len(day_timestamps)} timestamps encontrados")
            
            # Processar cada variável
            variables_processed = 0
            
            for variable in self.monthly_column_mapping.keys():
                # Verificar se a variável existe nos dados
                if variable not in available_variables:
                    print(f"⚠️  DEBUG: Variável {variable} não encontrada nos dados disponíveis")
                    continue
                
                # Coletar todos os valores do dia para esta variável
                day_values = []
                for ts in day_timestamps:
                    value = month_timestamps[ts].get(variable)
                    if value is not None:
                        day_values.append(value)
                
                if not day_values:
                    # Não há dados válidos para esta variável neste dia
                    print(f"⚠️  DEBUG: Variável {variable} - nenhum valor válido no dia {day}")
                    continue
                
                print(f"✅ DEBUG: Variável {variable} - {len(day_values)} valores no dia {day}")
                variables_processed += 1
                
                # Calcular estatísticas
                min_val = min(day_values)
                max_val = max(day_values)
                avg_val = sum(day_values) / len(day_values)
                
                # Calcular outliers
                outliers_count = self._calculate_outliers(day_values)
                
                # Obter posições das colunas
                col_info = self.monthly_column_mapping[variable]
                start_col = col_info['start_col']
                start_row, end_row = col_info['rows']
                
                # CORREÇÃO PRINCIPAL: Determinar qual linha usar baseado na seção da variável
                if start_row <= 33:  # Primeira seção (linhas 3-33)
                    target_row = day + 2  # dia 1 = linha 3, dia 2 = linha 4, etc.
                else:  # Segunda seção (linhas 37-67)
                    target_row = day + 36  # dia 1 = linha 37, dia 2 = linha 38, etc.
                
                # Verificar se estamos no range correto de linhas para esta variável
                if not (start_row <= target_row <= end_row):
                    print(f"❌ DEBUG: {variable} dia {day} - linha {target_row} fora do range {start_row}-{end_row}")
                    continue
                
                # Calcular letras das colunas (Min, Max, Avg, Outliers)
                from openpyxl.utils import column_index_from_string
                start_col_num = column_index_from_string(start_col)
                
                min_col = get_column_letter(start_col_num)      # Coluna Min
                max_col = get_column_letter(start_col_num + 1)  # Coluna Max  
                avg_col = get_column_letter(start_col_num + 2)  # Coluna Avg
                out_col = get_column_letter(start_col_num + 3)  # Coluna Outliers
                
                # Preencher células (usar ponto decimal, não vírgula)
                try:
                    ws[f'{min_col}{target_row}'] = round(min_val, 3)
                    ws[f'{max_col}{target_row}'] = round(max_val, 3)
                    ws[f'{avg_col}{target_row}'] = round(avg_val, 3)
                    ws[f'{out_col}{target_row}'] = int(outliers_count)
                    cells_updated += 4
                    print(f"✅ DEBUG: {variable} dia {day} - Min: {min_val:.3f}, Max: {max_val:.3f}, Avg: {avg_val:.3f}, Out: {outliers_count} (linha {target_row})")
                except Exception as e:
                    print(f"❌ DEBUG: Erro ao preencher {variable} dia {day} na linha {target_row}: {e}")
                    # Debug adicional - mostrar detalhes do erro
                    print(f"❌ DEBUG: Tentando escrever em {min_col}{target_row}, {max_col}{target_row}, {avg_col}{target_row}, {out_col}{target_row}")
                    pass
            
            print(f"🔍 DEBUG: Dia {day} - {variables_processed} variáveis processadas")
        
        print(f"🔍 DEBUG: Total de células atualizadas na análise mensal: {cells_updated}")
        return cells_updated

    def _debug_worksheet_structure(self, ws):
        """Debug da estrutura da planilha para entender o layout"""
        print(f"🔍 DEBUG: Analisando estrutura da aba {ws.title}")
        
        # Verificar algumas células chave
        test_cells = ['B3', 'B37', 'H37', 'N37', 'T37']
        
        for cell in test_cells:
            try:
                value = ws[cell].value
                print(f"  Célula {cell}: '{value}'")
            except Exception as e:
                print(f"  Célula {cell}: ERRO - {e}")
        
        # Verificar dimensões da planilha
        try:
            max_row = ws.max_row
            max_col = ws.max_column
            print(f"  Dimensões: {max_row} linhas x {max_col} colunas")
        except Exception as e:
            print(f"  Erro ao obter dimensões: {e}")

    def _debug_column_mapping(self):
        """Debug detalhado do mapeamento de colunas"""
        print("🔍 DEBUG: Verificando mapeamento de colunas mensais:")
        
        for variable, mapping in self.monthly_column_mapping.items():
            start_col = mapping['start_col']
            start_row, end_row = mapping['rows']
            
            from openpyxl.utils import column_index_from_string
            start_col_num = column_index_from_string(start_col)
            
            min_col = get_column_letter(start_col_num)
            max_col = get_column_letter(start_col_num + 1)
            avg_col = get_column_letter(start_col_num + 2)
            out_col = get_column_letter(start_col_num + 3)
            
            section = "PRIMEIRA" if start_row <= 33 else "SEGUNDA"
            
            print(f"  {variable}:")
            print(f"    Seção: {section}")
            print(f"    Linhas: {start_row} a {end_row}")
            print(f"    Colunas: {min_col}(Min) {max_col}(Max) {avg_col}(Avg) {out_col}(Out)")
            print(f"    Exemplo dia 1: linha {start_row if start_row <= 33 else 37}")
            print("")

    def _verify_data_variables(self, month_timestamps):
        """Verifica quais variáveis estão disponíveis nos dados"""
        if not month_timestamps:
            print("❌ DEBUG: Nenhum timestamp disponível")
            return []
        
        sample_data = next(iter(month_timestamps.values()))
        available_vars = list(sample_data.keys())
        mapped_vars = list(self.monthly_column_mapping.keys())
        
        print("🔍 DEBUG: Verificação de variáveis:")
        print(f"  Variáveis nos dados: {available_vars}")
        print(f"  Variáveis mapeadas: {mapped_vars}")
        
        missing_in_data = [var for var in mapped_vars if var not in available_vars]
        missing_in_mapping = [var for var in available_vars if var not in mapped_vars]
        
        if missing_in_data:
            print(f"  ❌ Variáveis mapeadas mas ausentes nos dados: {missing_in_data}")
        
        if missing_in_mapping:
            print(f"  ⚠️  Variáveis nos dados mas não mapeadas: {missing_in_mapping}")
        
        common_vars = [var for var in mapped_vars if var in available_vars]
        print(f"  ✅ Variáveis comuns (serão processadas): {common_vars}")
        
        return common_vars

    def _calculate_outliers(self, values):
        """Calcula número de outliers usando a fórmula padrão"""
        if len(values) < 2:
            return 0
        
        try:
            # Calcular quartis e IQR
            q1 = np.percentile(values, 25)
            q3 = np.percentile(values, 75)
            iqr = q3 - q1
            
            # Calcular média
            mean_val = np.mean(values)
            
            # Calcular limites
            l_sup = mean_val + 1.5 * iqr
            l_inf = mean_val - 1.5 * iqr
            
            # Contar outliers
            outliers = np.sum((np.array(values) < l_inf) | (np.array(values) > l_sup))
            
            return int(outliers)
        except Exception as e:
            print(f"❌ Erro no cálculo de outliers: {e}")
            return 0

    def _show_file_processing_summary(self):
        """Mostra resumo detalhado do processamento"""
        if hasattr(self, 'file_processing_info') and self.file_processing_info:
            st.markdown("---")
            st.markdown("### Resumo do Processamento")
            
            # Criar DataFrame com as informações
            df_files = pd.DataFrame(self.file_processing_info)
            
            # Calcular totais
            total_records = df_files['registros'].sum()
            total_files_success = len([f for f in self.file_processing_info if 'sucesso' in f['status'].lower()])
            total_timestamps = len(self.consolidated_data)
            
            # Mostrar métricas gerais
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.markdown(f"""
                <div class="metric-card">
                    <h4>Arquivos Processados</h4>
                    <h2>{total_files_success}</h2>
                </div>
                """, unsafe_allow_html=True)
            
            with col2:
                st.markdown(f"""
                <div class="metric-card">
                    <h4>Registros Lidos</h4>
                    <h2>{total_records:,}</h2>
                </div>
                """, unsafe_allow_html=True)
            
            with col3:
                st.markdown(f"""
                <div class="metric-card">
                    <h4>Timestamps Únicos</h4>
                    <h2>{total_timestamps:,}</h2>
                </div>
                """, unsafe_allow_html=True)
            
            with col4:
                st.markdown(f"""
                <div class="metric-card">
                    <h4>Conflitos Detectados</h4>
                    <h2>{len(self.conflicts_detected)}</h2>
                </div>
                """, unsafe_allow_html=True)
            
            # Tabela detalhada
            st.markdown("#### Detalhes por Arquivo")
            df_display = df_files.copy()
            df_display.columns = ['Arquivo', 'Registros', 'Início', 'Fim', 'Status']
            df_display['Registros'] = df_display['Registros'].apply(lambda x: f"{x:,}" if x > 0 else "0")
            
            st.dataframe(df_display, use_container_width=True)

    def get_updated_excel_file(self):
        """Retorna o arquivo Excel atualizado"""
        if self.excel_path and os.path.exists(self.excel_path):
            with open(self.excel_path, 'rb') as f:
                return f.read()
        return None

    def show_data_preview_and_charts(self):
        """Mostra preview dos dados consolidados com gráficos para conferência"""
        if not self.consolidated_data:
            return
        
        st.markdown("---")
        st.markdown("### Análise dos Dados Consolidados")
        
        # Converter para DataFrame para visualização
        preview_data = []
        for timestamp, data in self.consolidated_data.items():
            row = {'Timestamp': timestamp}
            row.update(data)
            preview_data.append(row)
        
        if preview_data:
            df_preview = pd.DataFrame(preview_data)
            df_preview = df_preview.sort_values('Timestamp')
            
            # Estatísticas gerais
            st.markdown("#### Estatísticas Gerais")
            col1, col2, col3 = st.columns(3)
            
            with col1:
                first_timestamp = min(self.consolidated_data.keys())
                last_timestamp = max(self.consolidated_data.keys())
                period_days = (last_timestamp - first_timestamp).days + 1
                st.metric("Período Total", f"{period_days} dias")
            
            with col2:
                timestamps_per_day = len(self.consolidated_data) / period_days if period_days > 0 else 0
                st.metric("Registros/Dia", f"{timestamps_per_day:.1f}")
            
            with col3:
                # Agrupar por mês
                months = set()
                for ts in self.consolidated_data.keys():
                    months.add(f"{ts.year}-{ts.month:02d}")
                st.metric("Meses Cobertos", len(months))
            
            # Gráficos para conferência das variáveis
            self._create_variable_charts(df_preview)
            
            # Preview da tabela de dados
            st.markdown("#### Preview dos Dados (Primeiros 100 registros)")
            st.dataframe(df_preview.head(100), use_container_width=True)
            
        else:
            st.info("Nenhum dado disponível para preview.")

    def _create_variable_charts(self, df):
        """Cria gráficos para conferência das variáveis meteorológicas"""
        st.markdown("#### Gráficos de Conferência das Variáveis")
        
        # Preparar dados para gráficos
        df_clean = df.dropna()
        
        if len(df_clean) == 0:
            st.warning("Não há dados suficientes para gerar gráficos.")
            return
        
        # Gráfico 1: Temperatura ao longo do tempo
        with st.container():
            st.markdown('<div class="chart-container">', unsafe_allow_html=True)
            st.markdown("**Temperatura (°C)**")
            
            fig_temp = px.line(df_clean, x='Timestamp', y='Temperatura',
                              title='Variação da Temperatura ao Longo do Tempo',
                              color_discrete_sequence=['#00529C'])
            fig_temp.update_layout(
                xaxis_title="Data/Hora",
                yaxis_title="Temperatura (°C)",
                height=400,
                showlegend=False
            )
            st.plotly_chart(fig_temp, use_container_width=True)
            st.markdown('</div>', unsafe_allow_html=True)
        
        # Gráfico 2: Piranômetros (Radiação Solar)
        with st.container():
            st.markdown('<div class="chart-container">', unsafe_allow_html=True)
            st.markdown("**Radiação Solar (kW/m²)**")
            
            fig_pir = go.Figure()
            
            if 'Piranometro_1' in df_clean.columns:
                fig_pir.add_trace(go.Scatter(x=df_clean['Timestamp'], y=df_clean['Piranometro_1'],
                                           mode='lines', name='Piranômetro 1', line=dict(color='#FF6B35')))
            
            if 'Piranometro_2' in df_clean.columns:
                fig_pir.add_trace(go.Scatter(x=df_clean['Timestamp'], y=df_clean['Piranometro_2'],
                                           mode='lines', name='Piranômetro 2', line=dict(color='#F7931E')))
            
            if 'Piranometro_Alab' in df_clean.columns:
                fig_pir.add_trace(go.Scatter(x=df_clean['Timestamp'], y=df_clean['Piranometro_Alab'],
                                           mode='lines', name='Piranômetro Alabiótico', line=dict(color='#FFD23F')))
            
            fig_pir.update_layout(
                title='Radiação Solar - Comparação dos Piranômetros',
                xaxis_title="Data/Hora",
                yaxis_title="Radiação Solar (kW/m²)",
                height=400,
                legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1)
            )
            st.plotly_chart(fig_pir, use_container_width=True)
            st.markdown('</div>', unsafe_allow_html=True)
        
        # Gráfico 3: Umidade e Velocidade do Vento
        with st.container():
            st.markdown('<div class="chart-container">', unsafe_allow_html=True)
            st.markdown("**Umidade Relativa e Velocidade do Vento**")
            
            # Criar subplots com eixos Y duplos
            fig_combined = make_subplots(specs=[[{"secondary_y": True}]])
            
            if 'Umidade_Relativa' in df_clean.columns:
                fig_combined.add_trace(
                    go.Scatter(x=df_clean['Timestamp'], y=df_clean['Umidade_Relativa'],
                             mode='lines', name='Umidade Relativa (%)', line=dict(color='#4A90E2')),
                    secondary_y=False,
                )
            
            if 'Velocidade_Vento' in df_clean.columns:
                fig_combined.add_trace(
                    go.Scatter(x=df_clean['Timestamp'], y=df_clean['Velocidade_Vento'],
                             mode='lines', name='Velocidade do Vento (m/s)', line=dict(color='#50C878')),
                    secondary_y=True,
                )
            
            # Configurar eixos Y
            fig_combined.update_xaxes(title_text="Data/Hora")
            fig_combined.update_yaxes(title_text="Umidade Relativa (%)", secondary_y=False)
            fig_combined.update_yaxes(title_text="Velocidade do Vento (m/s)", secondary_y=True)
            
            fig_combined.update_layout(
                title='Umidade Relativa e Velocidade do Vento',
                height=400,
                legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1)
            )
            st.plotly_chart(fig_combined, use_container_width=True)
            st.markdown('</div>', unsafe_allow_html=True)
        
        # Estatísticas descritivas
        st.markdown("#### Estatísticas Descritivas")
        
        # Selecionar apenas colunas numéricas
        numeric_cols = ['Temperatura', 'Piranometro_1', 'Piranometro_2', 'Piranometro_Alab', 'Umidade_Relativa', 'Velocidade_Vento']
        available_cols = [col for col in numeric_cols if col in df_clean.columns]
        
        if available_cols:
            stats_df = df_clean[available_cols].describe().round(3)
            stats_df.index = ['Contagem', 'Média', 'Desvio Padrão', 'Mínimo', '25%', '50% (Mediana)', '75%', 'Máximo']
            
            st.markdown('<div class="stats-container">', unsafe_allow_html=True)
            st.dataframe(stats_df, use_container_width=True)
            st.markdown('</div>', unsafe_allow_html=True)
        
        # Distribuição dos dados por hora do dia
        st.markdown("#### Distribuição por Hora do Dia")
        
        if len(df_clean) > 0:
            df_clean['Hora'] = df_clean['Timestamp'].dt.hour
            
            # Gráfico de distribuição por hora
            hourly_stats = df_clean.groupby('Hora')[available_cols].mean().reset_index()
            
            fig_hourly = go.Figure()
            
            colors = ['#00529C', '#FF6B35', '#F7931E', '#FFD23F', '#4A90E2', '#50C878']
            
            for i, col in enumerate(available_cols):
                if col in hourly_stats.columns:
                    fig_hourly.add_trace(go.Scatter(
                        x=hourly_stats['Hora'], 
                        y=hourly_stats[col],
                        mode='lines+markers',
                        name=col.replace('_', ' ').title(),
                        line=dict(color=colors[i % len(colors)])
                    ))
            
            fig_hourly.update_layout(
                title='Valores Médios por Hora do Dia',
                xaxis_title="Hora do Dia",
                yaxis_title="Valores Médios",
                height=400,
                legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1)
            )
            
            st.plotly_chart(fig_hourly, use_container_width=True)

    # === NOVO: FUNCIONALIDADES DO DASHBOARD ===
    
    def load_excel_data_for_dashboard(self):
        """Carrega dados do Excel atualizado para o dashboard"""
        if not self.excel_path or not os.path.exists(self.excel_path):
            return None
        
        try:
            wb = load_workbook(self.excel_path)
            excel_data = {}
            
            # Carregar dados das abas mensais
            for sheet_name in wb.sheetnames:
                if "Analise Mensal" in sheet_name:
                    # Extrair número do mês do nome da aba
                    month_str = sheet_name.split('-')[0]
                    try:
                        month_num = int(month_str)
                        excel_data[month_num] = self._read_monthly_sheet_data(wb[sheet_name], month_num)
                    except ValueError:
                        continue
            
            return excel_data
        except Exception as e:
            st.error(f"Erro ao carregar dados do Excel: {e}")
            return None

    def _read_monthly_sheet_data(self, ws, month_num):
        """Lê dados de uma aba de análise mensal"""
        monthly_data = {}
        
        # Para cada variável no mapeamento
        for variable, col_info in self.monthly_column_mapping.items():
            start_col = col_info['start_col']
            start_row, end_row = col_info['rows']
            
            from openpyxl.utils import column_index_from_string
            start_col_num = column_index_from_string(start_col)
            
            # Ler dados de cada dia
            variable_data = {
                'dias': [],
                'min': [],
                'max': [],
                'avg': [],
                'outliers': []
            }
            
            # Determinar range de linhas baseado na variável
            if start_row <= 33:  # Primeira seção
                day_range = range(3, 34)  # Linhas 3-33 (dias 1-31)
            else:  # Segunda seção
                day_range = range(37, 68)  # Linhas 37-67 (dias 1-31)
            
            for row_num in day_range:
                day = row_num - 2 if start_row <= 33 else row_num - 36
                
                # Verificar se o dia é válido para o mês
                try:
                    datetime(2024, month_num, day)  # Ano irrelevante para validação
                except ValueError:
                    continue
                
                # Ler valores das colunas Min, Max, Avg, Outliers
                try:
                    min_val = ws[f'{get_column_letter(start_col_num)}{row_num}'].value
                    max_val = ws[f'{get_column_letter(start_col_num + 1)}{row_num}'].value
                    avg_val = ws[f'{get_column_letter(start_col_num + 2)}{row_num}'].value
                    out_val = ws[f'{get_column_letter(start_col_num + 3)}{row_num}'].value
                    
                    # Só adicionar se houver pelo menos um valor válido
                    if any(v is not None for v in [min_val, max_val, avg_val, out_val]):
                        variable_data['dias'].append(day)
                        variable_data['min'].append(min_val if min_val is not None else 0)
                        variable_data['max'].append(max_val if max_val is not None else 0)
                        variable_data['avg'].append(avg_val if avg_val is not None else 0)
                        variable_data['outliers'].append(out_val if out_val is not None else 0)
                        
                except Exception:
                    continue
            
            if variable_data['dias']:  # Só adicionar se houver dados
                monthly_data[variable] = variable_data
        
        return monthly_data

    def show_dashboard(self):
        """Exibe o dashboard analítico"""
        st.markdown("---")
        st.markdown("""
        <div class="dashboard-header">
            <h2>📊 Dashboard de Análise Meteorológica</h2>
            <p>Visualização interativa dos dados processados</p>
        </div>
        """, unsafe_allow_html=True)
        
        # Carregar dados do Excel
        excel_data = self.load_excel_data_for_dashboard()
        
        if not excel_data and not self.consolidated_data:
            st.warning("Nenhum dado disponível para análise. Execute o processamento primeiro.")
            return
        
        # Filtros do Dashboard
        self._create_dashboard_filters(excel_data)
        
        # Aplicar filtros e mostrar visualizações
        self._show_filtered_analysis()

    def _create_dashboard_filters(self, excel_data):
        """Cria os filtros do dashboard"""
        st.markdown('<div class="dashboard-filters">', unsafe_allow_html=True)
        st.markdown("### 🔧 Filtros de Análise")
        
        col1, col2 = st.columns([2, 1])
        
        with col1:
            st.markdown("**Selecione os Meses para Análise:**")
            
            # Identificar meses disponíveis
            available_months = set()
            
            # Meses dos dados consolidados
            if self.consolidated_data:
                for timestamp in self.consolidated_data.keys():
                    available_months.add(timestamp.month)
            
            # Meses do Excel
            if excel_data:
                available_months.update(excel_data.keys())
            
            available_months = sorted(list(available_months))
            
            # Criar checkboxes para meses
            month_names = {
                1: 'JAN', 2: 'FEV', 3: 'MAR', 4: 'ABR', 
                5: 'MAI', 6: 'JUN', 7: 'JUL', 8: 'AGO',
                9: 'SET', 10: 'OUT', 11: 'NOV', 12: 'DEZ'
            }
            
            # Organizar checkboxes em colunas
            cols = st.columns(6)
            selected_months = []
            
            for i, month_num in enumerate(available_months):
                col_idx = i % 6
                with cols[col_idx]:
                    if st.checkbox(month_names[month_num], key=f"month_{month_num}"):
                        selected_months.append(month_num)
        
        with col2:
            st.markdown("**Tipo de Análise:**")
            analysis_type = st.radio(
                "Selecione o tipo:",
                ["Análise Diária", "Análise Mensal"],
                key="analysis_type"
            )
        
        # Salvar seleções no session_state
        st.session_state.selected_months = selected_months
        st.session_state.analysis_type = analysis_type
        
        st.markdown('</div>', unsafe_allow_html=True)
        
        # Mostrar resumo das seleções
        if selected_months:
            month_names_selected = [month_names[m] for m in selected_months]
            st.info(f"📅 **Meses selecionados:** {', '.join(month_names_selected)} | **Tipo:** {analysis_type}")
        else:
            st.warning("⚠️ Selecione pelo menos um mês para visualizar as análises.")

    def _show_filtered_analysis(self):
        """Mostra análises baseadas nos filtros selecionados"""
        if not hasattr(st.session_state, 'selected_months') or not st.session_state.selected_months:
            return
        
        selected_months = st.session_state.selected_months
        analysis_type = st.session_state.analysis_type
        
        if analysis_type == "Análise Diária":
            self._show_daily_analysis_dashboard(selected_months)
        else:
            self._show_monthly_analysis_dashboard(selected_months)

    def _show_daily_analysis_dashboard(self, selected_months):
        """Mostra dashboard de análise diária"""
        st.markdown("### 📈 Análise Diária - Dados Temporais")
        
        if not self.consolidated_data:
            st.warning("Dados consolidados não disponíveis para análise diária.")
            return
        
        # Filtrar dados pelos meses selecionados
        filtered_data = []
        for timestamp, data in self.consolidated_data.items():
            if timestamp.month in selected_months:
                row = {'Timestamp': timestamp}
                row.update(data)
                filtered_data.append(row)
        
        if not filtered_data:
            st.warning("Nenhum dado encontrado para os meses selecionados.")
            return
        
        df_filtered = pd.DataFrame(filtered_data)
        df_filtered = df_filtered.sort_values('Timestamp')
        
        # Estatísticas resumidas
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.metric("Total de Registros", f"{len(df_filtered):,}")
        with col2:
            period_days = (df_filtered['Timestamp'].max() - df_filtered['Timestamp'].min()).days + 1
            st.metric("Período (dias)", period_days)
        with col3:
            avg_records_per_day = len(df_filtered) / period_days if period_days > 0 else 0
            st.metric("Registros/Dia", f"{avg_records_per_day:.1f}")
        with col4:
            months_count = df_filtered['Timestamp'].dt.month.nunique()
            st.metric("Meses Analisados", months_count)
        
        # Gráficos combinados para múltiplos meses
        self._create_combined_daily_charts(df_filtered, selected_months)

    def _create_combined_daily_charts(self, df, selected_months):
        """Cria gráficos combinados para análise diária"""
        df_clean = df.dropna()
        
        if len(df_clean) == 0:
            st.warning("Não há dados suficientes para gerar gráficos.")
            return
        
        # Adicionar coluna de mês para colorir gráficos
        df_clean = df_clean.copy()
        df_clean['Mes'] = df_clean['Timestamp'].dt.month
        df_clean['Mes_Nome'] = df_clean['Mes'].map({
            1: 'JAN', 2: 'FEV', 3: 'MAR', 4: 'ABR', 
            5: 'MAI', 6: 'JUN', 7: 'JUL', 8: 'AGO',
            9: 'SET', 10: 'OUT', 11: 'NOV', 12: 'DEZ'
        })
        
        # Gráfico 1: Temperatura
        st.markdown('<div class="chart-container">', unsafe_allow_html=True)
        st.markdown("#### 🌡️ Temperatura ao Longo do Tempo")
        
        fig_temp = px.line(df_clean, x='Timestamp', y='Temperatura', 
                          color='Mes_Nome',
                          title='Variação da Temperatura por Mês',
                          labels={'Temperatura': 'Temperatura (°C)', 'Mes_Nome': 'Mês'})
        fig_temp.update_layout(height=500, showlegend=True)
        st.plotly_chart(fig_temp, use_container_width=True)
        st.markdown('</div>', unsafe_allow_html=True)
        
        # Gráfico 2: Radiação Solar (Piranômetros)
        st.markdown('<div class="chart-container">', unsafe_allow_html=True)
        st.markdown("#### ☀️ Radiação Solar - Piranômetros")
        
        fig_solar = go.Figure()
        
        colors = ['#FF6B35', '#F7931E', '#FFD23F']
        piranometer_vars = ['Piranometro_1', 'Piranometro_2', 'Piranometro_Alab']
        piranometer_names = ['Piranômetro 1', 'Piranômetro 2', 'Piranômetro Alabiótico']
        
        for i, (var, name) in enumerate(zip(piranometer_vars, piranometer_names)):
            if var in df_clean.columns:
                fig_solar.add_trace(go.Scatter(
                    x=df_clean['Timestamp'], 
                    y=df_clean[var],
                    mode='lines',
                    name=name,
                    line=dict(color=colors[i])
                ))
        
        fig_solar.update_layout(
            title='Radiação Solar - Comparação dos Sensores',
            xaxis_title="Data/Hora",
            yaxis_title="Radiação Solar (kW/m²)",
            height=500,
            showlegend=True
        )
        st.plotly_chart(fig_solar, use_container_width=True)
        st.markdown('</div>', unsafe_allow_html=True)
        
        # Gráfico 3: Umidade e Vento (Eixos duplos)
        st.markdown('<div class="chart-container">', unsafe_allow_html=True)
        st.markdown("#### 💨 Umidade Relativa e Velocidade do Vento")
        
        fig_env = make_subplots(specs=[[{"secondary_y": True}]])
        
        if 'Umidade_Relativa' in df_clean.columns:
            fig_env.add_trace(
                go.Scatter(x=df_clean['Timestamp'], y=df_clean['Umidade_Relativa'],
                         mode='lines', name='Umidade Relativa (%)', 
                         line=dict(color='#4A90E2')),
                secondary_y=False,
            )
        
        if 'Velocidade_Vento' in df_clean.columns:
            fig_env.add_trace(
                go.Scatter(x=df_clean['Timestamp'], y=df_clean['Velocidade_Vento'],
                         mode='lines', name='Velocidade do Vento (m/s)', 
                         line=dict(color='#50C878')),
                secondary_y=True,
            )
        
        fig_env.update_xaxes(title_text="Data/Hora")
        fig_env.update_yaxes(title_text="Umidade Relativa (%)", secondary_y=False)
        fig_env.update_yaxes(title_text="Velocidade do Vento (m/s)", secondary_y=True)
        fig_env.update_layout(
            title='Condições Ambientais - Umidade e Vento',
            height=500,
            showlegend=True
        )
        st.plotly_chart(fig_env, use_container_width=True)
        st.markdown('</div>', unsafe_allow_html=True)
        
        # Gráfico 4: Distribuição por hora do dia (todos os meses)
        st.markdown('<div class="chart-container">', unsafe_allow_html=True)
        st.markdown("#### ⏰ Padrões Diários - Médias por Hora")
        
        df_hourly = df_clean.copy()
        df_hourly['Hora'] = df_hourly['Timestamp'].dt.hour
        
        # Calcular médias por hora para cada mês
        hourly_stats = df_hourly.groupby(['Hora', 'Mes_Nome'])[['Temperatura', 'Umidade_Relativa', 'Velocidade_Vento']].mean().reset_index()
        
        fig_hourly = px.line(hourly_stats, x='Hora', y='Temperatura', 
                           color='Mes_Nome',
                           title='Temperatura Média por Hora do Dia (por Mês)',
                           labels={'Temperatura': 'Temperatura (°C)', 'Mes_Nome': 'Mês'})
        fig_hourly.update_layout(height=400)
        st.plotly_chart(fig_hourly, use_container_width=True)
        st.markdown('</div>', unsafe_allow_html=True)

    def _show_monthly_analysis_dashboard(self, selected_months):
        """Mostra dashboard de análise mensal"""
        st.markdown("### 📊 Análise Mensal - Estatísticas Diárias")
        
        # Carregar dados do Excel
        excel_data = self.load_excel_data_for_dashboard()
        
        if not excel_data:
            st.warning("Dados de análise mensal não disponíveis. Execute o processamento completo primeiro.")
            return
        
        # Filtrar dados pelos meses selecionados
        filtered_excel_data = {month: data for month, data in excel_data.items() if month in selected_months}
        
        if not filtered_excel_data:
            st.warning("Nenhum dado de análise mensal encontrado para os meses selecionados.")
            return
        
        # Estatísticas resumidas
        total_variables = 0
        total_days_with_data = 0
        
        for month_data in filtered_excel_data.values():
            total_variables += len(month_data)
            for var_data in month_data.values():
                total_days_with_data += len(var_data['dias'])
        
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.metric("Meses Analisados", len(filtered_excel_data))
        with col2:
            st.metric("Variáveis Processadas", total_variables)
        with col3:
            st.metric("Dias com Dados", total_days_with_data)
        with col4:
            avg_days_per_month = total_days_with_data / len(filtered_excel_data) if len(filtered_excel_data) > 0 else 0
            st.metric("Média Dias/Mês", f"{avg_days_per_month:.1f}")
        
        # Gráficos de barras para estatísticas mensais
        self._create_monthly_bar_charts(filtered_excel_data, selected_months)

    def _create_monthly_bar_charts(self, excel_data, selected_months):
        """Cria gráficos de barras para análise mensal"""
        month_names = {
            1: 'JAN', 2: 'FEV', 3: 'MAR', 4: 'ABR', 
            5: 'MAI', 6: 'JUN', 7: 'JUL', 8: 'AGO',
            9: 'SET', 10: 'OUT', 11: 'NOV', 12: 'DEZ'
        }
        
        # Variáveis principais para gráficos
        main_variables = ['Temperatura', 'Piranometro_1', 'Piranometro_2', 'Umidade_Relativa', 'Velocidade_Vento']
        
        for variable in main_variables:
            # Verificar se a variável existe nos dados
            var_data_available = any(variable in month_data for month_data in excel_data.values())
            
            if not var_data_available:
                continue
            
            st.markdown('<div class="chart-container">', unsafe_allow_html=True)
            st.markdown(f"#### 📊 {variable.replace('_', ' ').title()}")
            
            # Preparar dados para o gráfico
            chart_data = []
            
            for month_num in selected_months:
                if month_num not in excel_data or variable not in excel_data[month_num]:
                    continue
                
                var_data = excel_data[month_num][variable]
                month_name = month_names[month_num]
                
                # Calcular estatísticas mensais médias
                if var_data['avg']:
                    avg_min = np.mean(var_data['min']) if var_data['min'] else 0
                    avg_max = np.mean(var_data['max']) if var_data['max'] else 0
                    avg_avg = np.mean(var_data['avg']) if var_data['avg'] else 0
                    total_outliers = sum(var_data['outliers']) if var_data['outliers'] else 0
                    
                    chart_data.extend([
                        {'Mês': month_name, 'Estatística': 'Mínimo', 'Valor': avg_min},
                        {'Mês': month_name, 'Estatística': 'Máximo', 'Valor': avg_max},
                        {'Mês': month_name, 'Estatística': 'Média', 'Valor': avg_avg},
                        {'Mês': month_name, 'Estatística': 'Outliers', 'Valor': total_outliers}
                    ])
            
            if chart_data:
                df_chart = pd.DataFrame(chart_data)
                
                # Separar outliers dos outros valores (escalas diferentes)
                df_values = df_chart[df_chart['Estatística'] != 'Outliers']
                df_outliers = df_chart[df_chart['Estatística'] == 'Outliers']
                
                # Gráfico principal (Min, Max, Média)
                if not df_values.empty:
                    fig_main = px.bar(df_values, x='Mês', y='Valor', color='Estatística',
                                    title=f'{variable.replace("_", " ").title()} - Estatísticas Mensais',
                                    barmode='group',
                                    color_discrete_sequence=['#00529C', '#FF6B35', '#50C878'])
                    fig_main.update_layout(height=400)
                    st.plotly_chart(fig_main, use_container_width=True)
                
                # Gráfico de outliers (escala separada)
                if not df_outliers.empty:
                    fig_outliers = px.bar(df_outliers, x='Mês', y='Valor',
                                        title=f'{variable.replace("_", " ").title()} - Outliers Detectados',
                                        color_discrete_sequence=['#FF4444'])
                    fig_outliers.update_layout(height=300)
                    st.plotly_chart(fig_outliers, use_container_width=True)
            
            st.markdown('</div>', unsafe_allow_html=True)
        
        # Gráfico comparativo de outliers por variável
        st.markdown('<div class="chart-container">', unsafe_allow_html=True)
        st.markdown("#### 🎯 Comparativo de Outliers por Variável")
        
        outliers_summary = []
        for month_num in selected_months:
            if month_num not in excel_data:
                continue
                
            month_name = month_names[month_num]
            for variable in main_variables:
                if variable in excel_data[month_num]:
                    var_data = excel_data[month_num][variable]
                    total_outliers = sum(var_data['outliers']) if var_data['outliers'] else 0
                    outliers_summary.append({
                        'Mês': month_name,
                        'Variável': variable.replace('_', ' ').title(),
                        'Outliers': total_outliers
                    })
        
        if outliers_summary:
            df_outliers_summary = pd.DataFrame(outliers_summary)
            fig_outliers_comp = px.bar(df_outliers_summary, x='Variável', y='Outliers', color='Mês',
                                     title='Comparativo de Outliers por Variável e Mês',
                                     barmode='group')
            fig_outliers_comp.update_layout(height=400)
            st.plotly_chart(fig_outliers_comp, use_container_width=True)
        
        st.markdown('</div>', unsafe_allow_html=True)


def main():
    # Cabeçalho principal com logo da CSN
    st.markdown("""
    <div class="main-header">
        <div class="logo-container">
            <img src="https://upload.wikimedia.org/wikipedia/pt/e/eb/Companhia_Sider%C3%BArgica_Nacional.png" 
                 alt="Logo CSN" class="logo-img">
            <div class="header-text">
                <h1>Medições Usina Geradora Floriano</h1>
                <p>Processador de Dados - Medicoes Diarias/Mensais</p>
                <p><small>Busca Pontual | Tolerância ±10min | Zero Inferências</small></p>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    # Inicializar o processador
    if 'processor' not in st.session_state:
        st.session_state.processor = ExactWeatherProcessor()
    
    # Controle de exibição do dashboard
    if 'show_dashboard' not in st.session_state:
        st.session_state.show_dashboard = False
    
    # Verificar se deve mostrar dashboard
    if st.session_state.show_dashboard:
        # Botão para voltar ao processamento
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            if st.button("⬅️ Voltar ao Processamento", use_container_width=True):
                st.session_state.show_dashboard = False
                st.rerun()
        
        # Mostrar dashboard
        st.session_state.processor.show_dashboard()
        return
    
    # Interface principal de processamento
    # Sidebar com instruções
    with st.sidebar:
        st.markdown("### Instruções de Uso")
        st.markdown("""
        **Passo 1:** Upload do arquivo Excel anual
        
        **Passo 2:** Upload dos arquivos .dat (múltiplos)
        
        **Passo 3:** Clique em "Processar Dados"
        
        **Passo 4:** Baixe o Excel atualizado
        
        **Passo 5:** Acesse o Dashboard Analítico
        """)
        
        st.markdown("---")
        st.markdown("### Funcionalidades")
        st.markdown("""
        **Características:**
        - Busca pontual de dados (sem médias)
        - Tolerância de ±10 minutos
        - Zero inferências ou preenchimentos
        - Detecção de conflitos entre arquivos
        - Mapeamento preciso por timestamp
        - Gráficos de conferência das variáveis
        - **Análise mensal automática**
        - **🆕 Dashboard interativo com filtros**
        
        **Lógica de Busca:**
        - **Diária:** Para 10:00 → busca entre 09:50 e 10:10
        - **Mensal:** Agrega todos os dados do dia para estatísticas
        - Prioriza timestamp mais próximo
        - Deixa vazio se não há dados na tolerância
        """)
        
        st.markdown("---")
        st.markdown("### Dashboard Analítico")
        st.markdown("""
        **Filtros Disponíveis:**
        - ✅ Seleção múltipla de meses (JAN, FEV, etc.)
        - ✅ Tipo de análise (Diária vs Mensal)
        
        **Visualizações:**
        - 📈 Gráficos temporais (análise diária)
        - 📊 Gráficos de barras (análise mensal)
        - 🎯 Comparativos de outliers
        - ⏰ Padrões diários e sazonais
        """)
    
    # Layout principal
    col1, col2 = st.columns([1, 1])
    
    with col1:
        st.markdown("### Upload do Excel Anual")
        excel_file = st.file_uploader(
            "Selecione o arquivo Excel anual",
            type=['xlsx', 'xls'],
            help="Arquivo Excel com as abas XX-Analise Diaria e XX-Analise Mensal"
        )
    
    with col2:
        st.markdown("### Upload dos Arquivos .dat")
        dat_files = st.file_uploader(
            "Selecione os arquivos .dat (múltiplos)",
            type=['dat'],
            accept_multiple_files=True,
            help="Arquivos de dados meteorológicos (.dat) com timestamps de 10 em 10 minutos"
        )
    
    # Informações sobre os arquivos carregados
    if dat_files:
        st.markdown("### Arquivos .dat Carregados")
        files_info = []
        for file in dat_files:
            files_info.append({
                'Arquivo': file.name,
                'Tamanho': f"{file.size / 1024:.1f} KB"
            })
        df_files = pd.DataFrame(files_info)
        st.dataframe(df_files, use_container_width=True)
    
    # Botão de processamento
    if excel_file and dat_files:
        st.markdown("---")
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            if st.button("Processar Dados - Atualizar Excel", use_container_width=True):
                with st.spinner("Processando dados com busca pontual..."):
                    # Processar arquivos .dat
                    success = st.session_state.processor.process_dat_files(dat_files)
                    
                    if success:
                        st.success("Arquivos .dat processados e consolidados com sucesso!")
                        
                        # Mostrar preview dos dados com gráficos
                        st.session_state.processor.show_data_preview_and_charts()
                        
                        # Atualizar Excel
                        st.markdown("### Atualizando Excel ...")
                        excel_file.seek(0)  # Reset file pointer
                        success, message = st.session_state.processor.update_excel_file(excel_file)
                        
                        if success:
                            st.success(f"{message}")
                            
                            # Informações sobre abas atualizadas
                            if st.session_state.processor.processed_sheets:
                                st.markdown("### Abas Atualizadas")
                                for sheet in st.session_state.processor.processed_sheets:
                                    st.markdown(f"- {sheet}")
                            
                            # Botão de download
                            st.markdown("### Download do Arquivo Atualizado")
                            updated_excel = st.session_state.processor.get_updated_excel_file()
                            if updated_excel:
                                st.download_button(
                                    label="📥 Baixar Excel Atualizado",
                                    data=updated_excel,
                                    file_name=f"analise_completa_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    use_container_width=True
                                )
                            
                            # Botão do dashboard
                            st.markdown("### Dashboard de Análise")
                            if st.button("📊 Ver Dashboard de Análise", use_container_width=True):
                                st.session_state.show_dashboard = True
                                st.rerun()
                                    
                        else:
                            st.error(f"{message}")
                    else:
                        st.error("Erro ao processar arquivos .dat")
    
    # Informações adicionais
    if not excel_file or not dat_files:
        st.markdown("---")
        st.markdown("### Aguardando Arquivos")
        missing = []
        if not excel_file:
            missing.append("Arquivo Excel anual")
        if not dat_files:
            missing.append("Arquivos .dat")
        
        st.info(f"Por favor, faça upload dos seguintes arquivos: {', '.join(missing)}")
        
        if not dat_files:
            st.markdown("""
            **Sobre o Processamento Automático:**
            - **Análise Diária:** Busca pontual com tolerância ±10min
            - **Análise Mensal:** Estatísticas diárias automáticas (Min/Max/Avg/Outliers)
            - Detecta automaticamente tipo de aba (Diária vs Mensal)
            - Processa 9 variáveis meteorológicas
            - Não preenche dados que não existem
            - Gera gráficos para conferência visual dos dados
            - **🆕 Dashboard interativo para análise avançada**
            """)
    
    # Footer
    st.markdown("---")
    st.markdown("""
    <div style="text-align: center; color: #666; padding: 1rem;">
        <p>Processador de Dados Meteorológicos | Usina Geradora Floriano</p>
        <p><small>Versão 3.0 - Análises Diárias, Mensais e Dashboard Interativo</small></p>
    </div>
    """, unsafe_allow_html=True)


if __name__ == "__main__":
    main()
